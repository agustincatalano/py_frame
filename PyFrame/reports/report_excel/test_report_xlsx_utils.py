
from openpyxl.styles import Border,Fill
# from openpyxl.styles.borders import Borders
# from openpyxl.writer import Font, Color

def     __create_styles(wb, ws):
    fills={}
    fills['green']= __create_fill__('FF00FF00')
    fills['red']= __create_fill__('FFFF0000')
    fills['grey']=__create_fill__('FF777777')
    fills['yellow']= __create_fill__('FFEE11')
    style_pallette = {}
    style_pallette['green'] = ws.conditional_formatting.addDxfStyle(wb,None,None,fills['green'])
    style_pallette['red'] = ws.conditional_formatting.addDxfStyle(wb,None,None,fills['red'])
    style_pallette['grey'] = ws.conditional_formatting.addDxfStyle(wb,None,None,fills['grey'])
    style_pallette['yellow'] = ws.conditional_formatting.addDxfStyle(wb,None,None,fills['yellow'])

    return style_pallette, fills

def __create_borders():
    border = Border()
    border.border_style = Border.BORDER_THIN

    b = {}

    top_borders = Borders()
    top_borders.left = border
    top_borders.right= border
    top_borders.top = border

    bottom_borders = Borders()
    bottom_borders.bottom=border
    bottom_borders.left=border
    bottom_borders.right=border

    side_borders =  Borders()
    side_borders.right = border
    side_borders.left = border

    borders  = Borders()
    borders.left = border
    borders.top=border
    borders.right=border
    borders.bottom=border


    b['full'] = borders
    b['top'] = top_borders
    b['bottom'] = bottom_borders
    b['side'] = side_borders
    return b

def __create_fill__(ARGB_color_string):
    fill = Fill()
    fill.start_color.index=ARGB_color_string
    fill.end_color.index=ARGB_color_string
    fill.fill_type=Fill.FILL_SOLID
    return fill


def scenario_formatting(ws,cell,scenario_status,pallette):
    ws.conditional_formatting.addCustomRule(cell, {'type' : 'expression', 'dxfId' : pallette['green'], 'formula' : ['\'Test Execution Report\'!'+scenario_status+'="Passed"'], 'stopIfTrue' : '1'})
    ws.conditional_formatting.addCustomRule(cell, {'type' : 'expression', 'dxfId' : pallette['red'], 'formula' : ['\'Test Execution Report\'!'+scenario_status+'="Failed"'], 'stopIfTrue' : '1'})
    ws.conditional_formatting.addCustomRule(cell, {'type' : 'expression', 'dxfId' : pallette['grey'], 'formula' : ['\'Test Execution Report\'!'+scenario_status+'="Skipped"'], 'stopIfTrue' : '1'})

def __copy_template_table__(template,ws, row_index,table_name):
    title_row = -1
    row = 1
    while(title_row == -1 and row < 25):
        if(template.cell("A"+str(row)).value == table_name):
            title_row = row
        else:
            row += 1
    # Copy template
    style_attrs = ("borders", "font", "fill", "alignment", "number_format")
    columns = [chr(x) for x in range(ord('A'), ord('J'))]
    for row in template.range("A{0}:I{1}".format(title_row, title_row+2)):
        column_index = 0
        for cell in row:
            address = columns[column_index]+str(row_index)
            ws.cell(address).value = cell.value
            for attr in style_attrs:
                setattr(ws.cell(address).style, attr, getattr(cell.style, attr))
            column_index += 1
        row_index += 1
    row_index -= 1
    return row_index