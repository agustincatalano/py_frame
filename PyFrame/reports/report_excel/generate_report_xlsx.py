"""This module generate xlsx file with reports the ejecutions
"""
import os
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook, Workbook
from openpyxl.styles.fonts import Font
from openpyxl.styles import Color, Style
from openpyxl.styles import colors
from openpyxl.charts import (
    PieChart,
    Reference,
    Series
)
from configobj import ConfigObj

TRACEABILITY = 'Traceability'
END_COLUMN_UTIL = 9
EXECUTION_REPORT = "Test Execution Report"
EXECUTION_REPORT_SEL = "'%s'" % EXECUTION_REPORT + "!$%s"
EXECUTION_REPORT_FUN = "=+'%s'" % EXECUTION_REPORT + "!$%s"
REPORTS = '..\\reports'



CONFIG = ConfigObj(os.path.join(os.getcwd(), "..", "config", "config.cfg"))


def generate_report(output):
    """genereate report in format xlsx"""
    features = output["features"]
    # creamos lo que tiene el excel
    work_book, work_sheet, template = _instance_template()
    #recolecta los escenarios de la feature
    scenarios = sum([feature['scenarios'] for feature in features], [])
    #exportamos los esteps
    __export_to_xlsx_steps(work_book, features)
    #esportamos los graficos
    __export_to_xlsx_charts(work_book, scenarios)
    #guardar
    work_book.save(os.path.join(os.path.abspath(REPORTS), "results.xlsx"))


def _instance_template():
    """instance template for creacion of the exel file"""
    # Open template sheet
    template_file = os.path.join(os.path.dirname(os.path.realpath(__file__)),
                                 "user_report_template.xlsx")
    template_book = load_workbook(template_file)
    template = template_book.get_active_sheet()

    # Create sheet
    work_book = Workbook()
    work_sheet = work_book.get_active_sheet()

    #retornamos el excel vacio pero armado, estructura.
    return work_book, work_sheet, template


def __export_to_xlsx_steps(work_book, features):
    """stylborder son los bordes de la tablita"""

    styleborder = Style(font=Font(bold=False), border=Border(top=Side(border_style='thin', color=colors.BLACK),
                                                            left=Side(border_style='thin', color=colors.BLACK),
                                                            bottom=Side(border_style='thin', color=colors.BLACK),
                                                            right=Side(border_style='thin', color=colors.BLACK)))
    #tomamos los steps del .feature y demas
    steps = _gather_steps(features)
    #nos paramos en la pestana activa
    work_sheet = Workbook.get_active_sheet(work_book)
    # y le ponemos el titulo
    work_sheet.title = 'Execution Steps'
    #nos paramos en la primer celda
    row_index = 1

    work_sheet['A1'].style = styleborder
    work_sheet['A1'].value = 'Step'
    work_sheet['B1'].style = styleborder
    work_sheet['B1'].value = 'Ocurrencias'
    work_sheet['C1'].style = styleborder
    work_sheet['C1'].value = 'Ejecuciones'
    work_sheet['D1'].style = styleborder
    work_sheet['D1'].value = 'Tiempo promedio'
    work_sheet['E1'].style = styleborder
    work_sheet['E1'].value = 'Tiempo total'
    work_sheet['F1'].style = styleborder
    work_sheet['F1'].value = 'Scenarios'

    #por cada columna le agrega los datos y steps, etc
    for step in sorted(steps):
        cell = work_sheet.cell(row=row_index + 1, column=1)
        cell.value = step
        cell.style = styleborder
        cell.offset(column=1).value = steps[step]['appearances']
        cell.offset(column=1).style = styleborder
        cell.offset(column=2).value = steps[step]['quantity']
        cell.offset(column=2).style = styleborder
        cell.offset(column=3).value = '%.2fs' % \
                                      (steps[step]['total_duration'] /
                                       (steps[step]['quantity'] or 1))

        cell.offset(column=3).style = styleborder
        cell.offset(column=4).value = '%.2fs' % steps[step]['total_duration']
        # cell.offset(column=5).value = steps[step]['scenario']
        # cell.offset(column=5).style = styleborder
        cell.offset(column=4).style = styleborder
        if len(step) > work_sheet.column_dimensions['A'].width:
            work_sheet.column_dimensions['A'].width = len(step)
        work_sheet.column_dimensions['B'].width = 10
        work_sheet.column_dimensions['C'].width = 10
        work_sheet.column_dimensions['D'].width = 10
        work_sheet.column_dimensions['E'].width = 10
        # work_sheet.column_dimensions['F'].width = 20
        row_index += 1

        # le agrega filtro a la columna
        # work_sheet.auto_filter.ref = "F1:F" + str(row_index)


def __export_to_xlsx_charts(work_book, scenarios):
    """recorre los escenarios y creamos los graficos"""

    scenarios = [scenario for scenario in scenarios
                 if not any([tag in scenario.get('tags', [])
                             for tag in os.environ.get('SKIP_TAGS',
                                                       '').split(',')])]

    total_passed = sum(scenario['status'].count("passed")
                       for scenario in scenarios)

    total_failed = sum(scenario['status'].count("failed")
                       for scenario in scenarios)

    total_skipped = sum(scenario['status'].count("skipped")
                        for scenario in scenarios)

    #creamos la pestana y le damos le titulo
    work_sheet2 = Workbook.create_sheet(work_book)
    work_sheet2.title = "Metricas"

    work_sheet2.cell("O2").value = "Test automation rate:"
    work_sheet2.cell("O3").value = "Autoamatizados:"
    work_sheet2.cell("P3").value = total_passed + total_failed
    work_sheet2.cell("O4").value = "Ignorados:"
    work_sheet2.cell("P4").value = total_skipped
    #rango de referencia de donde toma los valores el grafico
    values = Reference(work_sheet2, (3, 16), pos2=(4, 16))
    labels = Reference(work_sheet2, (3, 15), pos2=(4, 15))
    __add_chart(work_sheet2, values, labels, "Test Automation Rate", top=60)

    work_sheet2.cell("O6").value = "Test execution status"
    work_sheet2.cell("P7").value = total_passed
    work_sheet2.cell("P8").value = total_failed
    work_sheet2.cell("P9").value = total_skipped
    work_sheet2.cell("O7").value = "Passed"
    work_sheet2.cell("O8").value = "Failed"
    work_sheet2.cell("O9").value = "Skipped"
    values = Reference(work_sheet2, (7, 16), pos2=(9, 16))
    labels = Reference(work_sheet2, (7, 15), pos2=(9, 15))
    __add_chart(work_sheet2, values, labels, "Test Automation Rate", top=500)


def __add_chart(work_sheet, values, labels, title, top=400):
    """add chart"""
    series = Series(values, title=title, labels=labels, color=Color(colors.GREEN))
    chart = PieChart()
    chart.append(series)
    chart.drawing.top = top
    chart.drawing.left = 10
    work_sheet.add_chart(chart)


def _gather_steps(features):
    """retrieve a dictionary with steps used in latest test execution,
    together with the number of times they were executed"""
    steps = {}
    for feature in features:
        for scenario in feature['scenarios']:
            for step in scenario['steps']:
                if not step['name'] in steps:
                    steps[step['name']] = {'quantity': 0, 'total_duration': 0,
                                           'appearances': 0}
                steps[step['name']]['appearances'] += 1
                add = 1 if step['status'] != 'skipped' else 0
                steps[step['name']]['quantity'] += add
                steps[step['name']]['total_duration'] += step['duration']

    return steps
