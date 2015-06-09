import os
import json
from configobj import ConfigObj

from openpyxl import load_workbook, Workbook
# from openpyxl.writer import charts
import openpyxl.cell as Cell


import patched_charts
# from openpyxl.style import Font, Color
# from openpyxl.chart import PieChart, Reference, Series

from test_report_xlsx_utils import __create_styles, \
    __create_borders, scenario_formatting, __copy_template_table__


# FWK_DIR = os.environ.get('YARARA_PATH')
# charts.PieChartWriter = patched_charts.PieChartWriter
INFO_FILE = "test_results.json"
CONFIG = ConfigObj(os.path.join(os.getcwd(), "..", "config", "config.cfg"))
REPORTS = '..\\reports\\'
def generate_report():
    # Get info
    info_file = os.path.join(
        os.path.abspath(REPORTS), INFO_FILE)
    f = open(info_file, 'r')
    jsonString = f.read()
    f.close()
    output = json.loads(jsonString)
    features = output["features"]
    # Open template sheet
    cur_dir = os.path.dirname(os.path.realpath(__file__))
    template_file = os.path.join(cur_dir, "user_report_template.xlsx")
    template_book = load_workbook(template_file)
    template = template_book.get_active_sheet()
    # Create sheet
    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Test Execution Report"
    # Copy dimensions
    ws.row_dimensions = template.row_dimensions
    ws.column_dimensions = template.column_dimensions
    # Set first row
    global row_index
    row_index = 1
    # Fill test execution info
    row_index = __export_to_xlsx_general_summary(
        template, ws, features, row_index)
    row_index += 2
    row_index = __export_to_xlsx_feature_summary(
        template, ws, features, row_index)
    row_index += 2
    global start_index, pallette
    start_index = row_index + 2
    row_index, scenario_status_locator = __export_to_xlsx_test_cases(
        template, ws, features, row_index)

    pallette, fills = __create_styles(wb, ws)
    ws.conditional_formatting.addCustomRule("A" + str(start_index) + ":I" +
                                            str(row_index),
                                            {'type': 'expression',
                                             'dxfId': pallette['green'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Passed"'], 'stopIfTrue':
                                                '1'})

    ws.conditional_formatting.addCustomRule("A" + str(start_index) + ":I" +
                                            str(row_index),
                                            {'type': 'expression',
                                             'dxfId': pallette['red'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Failed"'], 'stopIfTrue':
                                                '1'})
    ws.conditional_formatting.addCustomRule("A" + str(start_index) + ":I" +
                                            str(row_index),
                                            {'type': 'expression',
                                             'dxfId': pallette['grey'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Skipped"'], 'stopIfTrue':
                                                '1'})
    # Save file
    __export_to_xlsx_traceability(wb, features, scenario_status_locator)
    __export_to_xlsx_tag_mapping(wb, features, scenario_status_locator)
    __export_to_xlsx_steps(wb, features)
    __export_to_xlsx_charts(wb, ws, features)
    wb.save(
        os.path.join(os.path.abspath(os.environ.get('OUTPUT')),
                     'test_results.xlsx'))


def __export_to_xlsx_steps(wb, features):
    steps = gather_steps(features)
    ws = Workbook.create_sheet(wb)
    ws.title = 'Steps'
    row_index = 1
    borders = __create_borders()
    ws['B1'].value = 'Step'
    ws['B1'].style.borders = borders['full']
    ws['C1'].value = 'Times used'
    ws['C1'].style.borders = borders['full']
    for step in sorted(steps):
        cell = ws.cell(row=row_index, column=1)
        cell.value = step
        cell.style.borders = borders['full']
        cell.offset(column=1).value = steps[step]
        cell.offset(column=1).style.borders = borders['full']
        if len(step) > ws.column_dimensions['B'].width:
            ws.column_dimensions['B'].width = len(step)
        row_index += 1
    ws.auto_filter = "B1:B" + str(row_index)


def __export_to_xlsx_traceability(wb, features, scenario_status_locator):
    story_tag_regex = CONFIG["story_tag"]["regex"]
    story_tag_prefix = CONFIG["story_tag"]["prefix"]
    stories, storyless_scenarios = gather_stories(
        features, story_tag_regex, story_tag_prefix)
    ws = Workbook.create_sheet(wb)
    ws.title = 'Traceability'
    ws['B1'].value = 'Story/bug'
    ws.conditional_formatting.addCustomRule("B2" + ":C" +
                                            str(row_index + 1),
                                            {'type': 'expression',
                                             'dxfId': pallette['green'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Passed"'], 'stopIfTrue':
                                                '1'})
    ws.conditional_formatting.addCustomRule("B2" + ":C" +
                                            str(row_index + 1),
                                            {'type': 'expression',
                                             'dxfId': pallette['red'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Failed"'], 'stopIfTrue':
                                                '1'})
    ws.conditional_formatting.addCustomRule("B2" + ":C" +
                                            str(row_index + 1),
                                            {'type': 'expression',
                                             'dxfId': pallette['grey'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Skipped"'], 'stopIfTrue':
                                                '1'})
    ws.conditional_formatting.addCustomRule("B2" + ":B" +
                                            str(row_index + 1),
                                            {'type': 'expression',
                                             'dxfId': pallette['green'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Passed"'], 'stopIfTrue':
                                                '1'})
    ws.conditional_formatting.addCustomRule("B2" + ":B" +
                                            str(row_index + 1),
                                            {'type': 'expression',
                                             'dxfId': pallette['red'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Failed"'], 'stopIfTrue':
                                                '1'})
    ws.conditional_formatting.addCustomRule("B2" + ":B" +
                                            str(row_index + 1),
                                            {'type': 'expression',
                                             'dxfId': pallette['grey'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Skipped"'], 'stopIfTrue':
                                                '1'})

    # Save file
    __create_xlsx_tag_filter(
        wb, ws, scenario_status_locator, stories, storyless_scenarios)


def __export_to_xlsx_tag_mapping(wb, features, scenario_status_locator):
    story_tag_regex = CONFIG["story_tag"]["regex"]
    story_tag_prefix = CONFIG["story_tag"]["prefix"]
    tags, tagless_scenarios = gather_stories(
        features, '(?!' + story_tag_regex + ')', story_tag_prefix)
    ws = Workbook.create_sheet(wb)
    ws.title = 'Tag Mapping'
    ws['B1'].value = 'Tag'
    ws.conditional_formatting.addCustomRule("B2" + ":C" +
                                            str(row_index + 1),
                                            {'type': 'expression',
                                             'dxfId': pallette['green'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Passed"'], 'stopIfTrue':
                                                '1'})
    ws.conditional_formatting.addCustomRule("B2" + ":C" +
                                            str(row_index + 1),
                                            {'type': 'expression',
                                             'dxfId': pallette['red'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Failed"'], 'stopIfTrue':
                                                '1'})
    ws.conditional_formatting.addCustomRule("B2" + ":C" +
                                            str(row_index + 1),
                                            {'type': 'expression',
                                             'dxfId': pallette['grey'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Skipped"'], 'stopIfTrue':
                                                '1'})
    ws.conditional_formatting.addCustomRule("B2" + ":B" +
                                            str(row_index + 1),
                                            {'type': 'expression',
                                             'dxfId': pallette['grey'],
                                             'formula':
                                                ['\'Test Execution Report\'!$G'
                                                 + str(start_index) +
                                                 '="Skipped"'], 'stopIfTrue':
                                                '1'})
    __create_xlsx_tag_filter(
        wb, ws, scenario_status_locator, tags, tagless_scenarios)


def __create_xlsx_tag_filter(wb, ws2, scenario_status_locator,
                             tags, tagless_scenarios):
    style_pallette, fills = __create_styles(wb, ws2)
    fills = {'failed': fills['red'], 'passed': fills[
        'green'], 'skipped': fills['grey'], 'yellow': fills['yellow']}
    #font_colors = {
    #    'failed': Color.RED, 'passed': Color.GREEN, 'skipped': "00777777"}
    borders = __create_borders()
    #font = Font()
    row_index = 1
    row_index_s = 1
    ws2['B1'].style.borders = borders['full']
    ws2.column_dimensions['B'].width = len(ws2['B1'].value) + 5
    ws2['C1'].value = 'Scenario'
    ws2['C1'].style.borders = borders['full']
    for tag in tags:
        tag_status = ''
        for scenario in tags[tag]:
            if tag_status != 'failed':
                if tag_status != 'skipped':
                    tag_status = scenario['status']
                elif scenario['status'] == 'failed':
                    tag_status = 'failed'
        for scenario in tags[tag]:
            cell = ws2.cell(row=row_index, column=2)
            cell.value = scenario['name']
            cell.style.borders = borders['full']
            scenario_formatting(
                ws2, 'C' + str(row_index + 1),
                scenario_status_locator[scenario['name']], style_pallette)
            if len(scenario['name']) > ws2.column_dimensions['C'].width:
                ws2.column_dimensions['C'].width = len(scenario['name'])
            cell.offset(column=-1).value = tag
            cell.offset(column=-1).style.fill = fills[tag_status]
            cell.offset(
                column=-1).style.font.color.index = Color.BLACK
            cell.offset(column=-1).style.borders = borders['full']
            row_index += 1
        cell = ws2.cell(row=row_index_s, column=1)
        cell.style.font.color.index = Color.BLACK
        cell.style.borders = borders['full']
        row_index_s += len(tags[tag])
        cell = ws2.cell(row=row_index_s - 1, column=1)
        cell.style.borders = borders['full']
        if len(tag) > ws2.column_dimensions['B'].width:
            ws2.column_dimensions['B'].width = len(tag) + 5

    for scenario in tagless_scenarios:
        cell = ws2.cell(row=row_index, column=2)
        cell.value = scenario['name']
        cell.style.borders = borders['full']
        scenario_formatting(ws2, 'C' + str(row_index + 1),
                            scenario_status_locator[scenario['name']],
                            style_pallette)
        if len(scenario['name']) > ws2.column_dimensions['C'].width:
            ws2.column_dimensions['C'].width = len(scenario['name'])

        cell.offset(column=-1).style.borders = borders['full']
        row_index += 1

    #ws2.cell.offset(column=-1).style.fill = fills['yellow']

    ws2.auto_filter = "B1:C1"  # + str(row_index)
    ws2['E2'].value = ws2['B1'].value + ' Legend: '
    ws2['E3'].value = 'All scenarios passed'
    ws2['E3'].style.fill = fills['passed']
    ws2['E4'].value = 'At least one scenario failed'
    ws2['E4'].style.fill = fills['failed']
    ws2['E5'].value = 'At least one scenario skipped'
    ws2['E5'].style.fill = fills['skipped']
    ws2['E6'].value = ws2['B1'].value + \
        ' column wont update' \
        ' automatically when test status changes'

    ws2.column_dimensions['E'].width = len(ws2['E6'].value)


def __export_to_xlsx_general_summary(template, ws, features, row_index):
    fields = {"features": 1, "number": 2, "executed": 3, "passed": 4,
              "failed": 5, "skipped": 6, "status": 7, "rate": 8, "duration": 9}

    row_index = __copy_template_table__(template, ws, row_index, "Summary")
    model_row_index = row_index

    values = dict()
    for field in fields.keys():
        values[field] = 0

    # Fill values
    first_test_index = 12 + len(features)
    total_scenarios = sum(len(feature['scenarios']) for feature in features)
    status_calc_range = 'G' + \
                        str(first_test_index) + ':G' + \
                        str(first_test_index + total_scenarios - 1)

    values['features'] = len(features)
    # sum(len(feature['scenarios']) for feature in features)
    values['number'] = total_scenarios
    values['passed'] = '=COUNTIF(' + status_calc_range + ',"Passed")'
    values['failed'] = '=COUNTIF(' + status_calc_range + ',"Failed")'
    values['executed'] = '=COUNTIF(' + status_calc_range + \
        ',"Failed") + COUNTIF(' + status_calc_range + ',"Passed")'
    values['skipped'] = '=COUNTIF(' + status_calc_range + ',"Skipped")'
    values['rate'] = '=COUNTIF(' + status_calc_range + \
                     ',"Passed")/' + str(total_scenarios) + '*100'
    values['duration'] = sum(feature['duration'] for feature in features)
    values['status'] = '=COUNTIF(' + status_calc_range + \
                       ',"<>Skipped")/' + str(total_scenarios) + '*100'
    # Fill info
    for field in fields.keys():
        model_cell = ws.cell(row=model_row_index - 1, column=fields[field] - 1)
        cell = ws.cell(row=row_index - 1, column=fields[field] - 1)
        cell.value = values[field]
        # Advance row
    row_index += 1
    # Return next free row
    return row_index


def __export_to_xlsx_feature_summary(template, ws, features, row_index):
    fields = {"feature": 1, "number": 2, "executed": 3, "passed": 4,
              "failed": 5, "skipped": 6, "status": 7, "rate": 8, "duration": 9}
    row_index = __copy_template_table__(template, ws, row_index, "Summary")

    last_test_index = row_index + 4 + len(features)
    model_row_index = row_index
    style_attrs = ("borders", "font", "fill", "alignment", "number_format")
    # Fill values
    for feature in features:
        # Get info
        values = dict()
        for field in fields.keys():
            values[field] = 0
        values['feature'] = feature['name']
        values['number'] = len(feature['scenarios'])

        status_calc_range = 'G' + \
            str(last_test_index) + ':G' + \
            str(last_test_index + len(feature['scenarios']) - 1)
        last_test_index += len(feature['scenarios'])

        values['status'] = '=COUNTIF(' + status_calc_range + \
            ',"<>Skipped")/' + str(len(feature['scenarios'])) + '*100'
        values['passed'] = '=COUNTIF(' + status_calc_range + ',"Passed")'
        values['failed'] = '=COUNTIF(' + status_calc_range + ',"Failed")'
        values['executed'] = '=COUNTIF(' + status_calc_range + \
            ',"Failed") + COUNTIF(' + status_calc_range + ',"Passed")'
        values['skipped'] = '=COUNTIF(' + status_calc_range + ',"Skipped")'
        values['rate'] = '=COUNTIF(' + status_calc_range + \
            ',"Passed")/' + str(len(feature['scenarios'])) + '*100'
        values['duration'] = feature['duration']
        # Fill info
        for field in fields.keys():
            model_cell = ws.cell(
                row=model_row_index - 1, column=fields[field] - 1)
            cell = ws.cell(row=row_index - 1, column=fields[field] - 1)
            for attr in style_attrs:
                setattr(cell.style, attr, getattr(model_cell.style, attr))
            cell.value = values[field]
        # Advance row
        row_index += 1
    # Return next free row
    return row_index


def __export_to_xlsx_test_cases(template, ws, features, row_index):
    scenario_status_locator = {}
    style_attrs = ("borders", "font", "fill", "alignment", "number_format")
    fields = {"feature": 1, "scenario": 2, "steps": 3, "empty1": 4,
              "empty2": 5, "automatic": 6, "status": 7, "duration": 8,
              "tester": 9}
    row_index = __copy_template_table__(template, ws, row_index, "Test Cases")
    model_row_index = row_index
    # Fill values
    for feature in features:
        # Get info
        values = dict()
        for field in fields.keys():
            values[field] = 0
        values['feature'] = feature['name']
        # Get first automatic cases
        scenarios_auto = []
        scenarios_manual = []
        for scenario in feature['scenarios']:
            if ("MANUAL" in scenario['tags']):
                scenarios_manual.append(scenario)
            else:
                scenarios_auto.append(scenario)
        scenarios = scenarios_auto + scenarios_manual
        for scenario in scenarios:
            values['scenario'] = scenario['name']
            values['status'] = scenario['status'].capitalize()
            values['duration'] = scenario['duration']
            if ("MANUAL" in scenario['tags']):
                values['automatic'] = "Manual"
                values['tester'] = ''
            else:
                values['automatic'] = "Automatic"
                values['tester'] = "Automatic"
            values['steps'] = ''
            for step in scenario['steps']:
                fail_string = ""
                if step['status'] == "failed":
                    error_msg, error_lines = gather_errors(scenario)
                    fail_string = " (FAILED!)\n"
                    if error_msg:
                        fail_string += "-" + error_msg + "\n"
                        for line in error_lines:
                            fail_string = fail_string + "-" + line + "\n"
                values['steps'] = values[
                    'steps'] + step['step_type'].capitalize() + ' ' \
                    + step['name'] + fail_string + '\n'
                if 'table' in step:
                    table_length = 0
                    for key in step['table']:
                        values['steps'] = values['steps'] + '|' + key
                        table_length = len(step['table'][key])
                    values['steps'] += '|\n'
                    for i in range(table_length):
                        for key in step['table']:
                            values['steps'] += '|' + step['table'][key][i]
                        values['steps'] += '|\n'

            # Fill info
            for field in fields.keys():
                model_cell = ws.cell(
                    row=model_row_index - 1, column=fields[field] - 1)
                cell = ws.cell(row=row_index - 1, column=fields[field] - 1)
                if field == 'status':
                    scenario_status_locator[values['scenario']] = \
                        Cell.absolute_coordinate(cell.get_coordinate())
                for attr in style_attrs:
                    setattr(cell.style, attr, getattr(model_cell.style, attr))
                cell.value = values[field]
            # Advance row
            row_index += 1
    # Return next free row
    return row_index, scenario_status_locator


def __export_to_xlsx_charts(wb, ws, features):
    total_passed = sum(sum(scenario['status'].count(
        "passed") for scenario in feature['scenarios'])
        for feature in features)
    total_failed = sum(sum(scenario['status'].count(
        "failed") for scenario in feature['scenarios'])
        for feature in features)
    total_skipped = sum(sum(scenario['status'].count(
        "skipped") for scenario in feature['scenarios'])
        for feature in features)
    ws2 = Workbook.create_sheet(wb)
    ws2.title = "Summary report"

    ws2.cell(
        "A1").value = "Charts are not dinamically updated.\
    If test status changes in test execution please\
    update the summarized value"

    ws2.cell("O2").value = "Test automation Rate"
    ws2.cell("P3").value = total_passed + total_failed
    ws2.cell("P4").value = total_skipped
    ws2.cell("O3").value = "Automated"
    ws2.cell("O4").value = "Not Automated"
    values = Reference(ws2, (2, 15), pos2=(3, 15))
    labels = Reference(ws2, (2, 14), pos2=(3, 14))
    __add_chart(ws2, values, labels, "Test automation Rate", top=60)

    ws2.cell("O6").value = "Test execution status"
    ws2.cell("P7").value = total_passed
    ws2.cell("P8").value = total_failed
    ws2.cell("P9").value = total_skipped
    ws2.cell("O7").value = "Passed"
    ws2.cell("O8").value = "Failed"
    ws2.cell("O9").value = "Skipped"
    values = Reference(ws2, (6, 15), pos2=(8, 15))
    labels = Reference(ws2, (6, 14), pos2=(8, 14))
    __add_chart(ws2, values, labels, "Test execution status", top=460)

    ws2.cell("O11").value = "Coverage"
    ws2.cell("P12").value = 0
    ws2.cell("P13").value = 0
    ws2.cell("O12").value = "Covered"
    ws2.cell("O13").value = "Uncovered"
    values = Reference(ws2, (11, 15), pos2=(12, 15))
    labels = Reference(ws2, (11, 14), pos2=(12, 14))
    __add_chart(ws2, values, labels, "Story Coverage", top=860)

    # Executed
    ws2.cell("P3").value = "=+'Test Execution Report'!C3"

    # Passed
    ws2.cell("P7").value = "=+'Test Execution Report'!D3"

    # Failed
    ws2.cell("P8").value = "=+'Test Execution Report'!E3"

    # Skipped
    ws2.cell("P9").value = "=+'Test Execution Report'!F3"


def __add_chart(ws, values, labels, title, left=10, top=400):
    series = Series(values, title=title, labels=labels, color=Color.GREEN)
    series.color = Color.BLACK
    chart = PieChart()
    chart.append(series)
    chart.drawing.top = top
    chart.drawing.left = left
    ws.add_chart(chart)


if __name__ == "__main__":
    # Make sure we are in the common run path (this script should by in
    # (root)/utils/, run path should be (root)/)
    script_dir, _ = os.path.split(os.path.abspath(__file__))
    run_dir, _ = os.path.split(script_dir)
    os.chdir(run_dir)
    # Call exports function
