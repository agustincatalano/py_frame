"""Helpers for xlsx"""
from openpyxl.styles import Fill, Border, borders, PatternFill
from openpyxl.formatting import ConditionalFormatting


def __create_styles(work_book, work_sheet):
    """Create style for xlsx file"""
    fills = {}
    fills['green'] = __create_fill__('FF00FF00')
    fills['red'] = __create_fill__('FFFF0000')
    fills['grey'] = __create_fill__('FF777777')
    fills['yellow'] = __create_fill__('FFEE11')
    style_pallette = {}
    style_pallette['green'] = work_sheet.conditional_formatting.setDxfStyles(
         work_book, fills['green'])
    style_pallette['red'] = work_sheet.conditional_formatting.setDxfStyles(
        work_book, None, None, fills['red'])
    style_pallette['grey'] = work_sheet.conditional_formatting.setDxfStyles(
        work_book, None, None, fills['grey'])
    style_pallette['yellow'] = work_sheet.conditional_formatting.setDxfStyles(
        work_book, None, None, fills['yellow'])

    return style_pallette, fills


def __create_borders():
    """Create dictionary of the borders """
    border = Border()
    border.border_style = Border.BORDER_THIN

    dict_bordors = {}

    top_borders = Border()
    top_borders.left = border
    top_borders.right = border
    top_borders.top = border

    bottom_borders = Border()
    bottom_borders.bottom = border
    bottom_borders.left = border
    bottom_borders.right = border

    side_borders = Border()
    side_borders.right = border
    side_borders.left = border

    borders = Border()
    borders.left = border
    borders.top = border
    borders.right = border
    borders.bottom = border


    dict_bordors['full'] = borders
    dict_bordors['top'] = top_borders
    dict_bordors['bottom'] = bottom_borders
    dict_bordors['side'] = side_borders
    return dict_bordors


def __create_fill__(argb_color_string):
    """Create fill"""

    fill = PatternFill(fill_type='solid',
                start_color='FFFFFFFF',
               end_color='FF000000')

    # fill.start_color.index = argb_color_string
    # fill.end_color.index = argb_color_string
    # fill.fill_type =
    return fill


def scenario_formatting(work_sheet, cell, scenario_status, pallette):
    """formattomg for scenario"""
    work_sheet.conditional_formatting.\
        addCustomRule(cell, {'type': 'expression',
                             'dxfId': pallette['green'],
                             'formula': ['\'Test Execution Report\'!' +
                                         scenario_status + '="Passed"'],
                             'stopIfTrue': '1'})

    work_sheet.conditional_formatting.\
        addCustomRule(cell, {'type': 'expression',
                             'dxfId': pallette['red'],
                             'formula': ['\'Test Execution Report\'!' +
                                         scenario_status + '="Failed"'],
                             'stopIfTrue': '1'})

    work_sheet.conditional_formatting.\
        addCustomRule(cell, {'type': 'expression',
                             'dxfId': pallette['grey'],
                             'formula': ['\'Test Execution Report\'!' +
                                         scenario_status + '="Skipped"'],
                             'stopIfTrue': '1'})


def __copy_template_table__(template, work_sheet, row_index, table_name):
    title_row = -1
    row = 1
    while title_row == -1 and row < 25:
        if template.cell("A" + str(row)).value == table_name:
            title_row = row
        else:
            row += 1
    # Copy template
    # style_attrs = ("borders", "font", "fill", "alignment", "number_format")
    style_attrs = ("font", "fill", "alignment", "number_format")
    columns = [chr(x) for x in range(ord('A'), ord('J'))]
    for row in template.range("A{0}:I{1}".format(title_row, title_row + 2)):
        column_index = 0
        for cell in row:
            address = columns[column_index] + str(row_index)
            work_sheet.cell(address).value = cell.value
            for attr in style_attrs:
                setattr(work_sheet.cell(address).style, attr,
                        getattr(cell.style, attr))
            column_index += 1
        row_index += 1
    row_index += 1
    return row_index
